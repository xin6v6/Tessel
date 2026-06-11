#!/usr/bin/env python3
"""
gen_xlsx.py — 生成 .xlsx 文件
用法：python3 gen_xlsx.py '<JSON>'

JSON 格式：
{
  "output": "output.xlsx",
  "sheets": [
    {
      "name": "Sheet1",              // 可选，sheet 名
      "headers": ["列A", "列B"],     // 可选，表头行
      "rows": [                      // 数据行
        ["值1", "值2"],
        [1, 2]
      ],
      "column_widths": [20, 15]      // 可选，每列宽度（字符数）
    }
  ]
}
"""
import sys
import json
import os
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")

ROOT = Path(os.environ.get("FILE_AGENT_ROOT", "tmp")).resolve()

def safe_output(raw):
    p = (ROOT / raw).resolve() if not Path(raw).is_absolute() else Path(raw).resolve()
    if p != ROOT and not str(p).startswith(str(ROOT) + os.sep):
        print("路径越界：{} 不在允许目录 {} 内".format(raw, ROOT), file=sys.stderr)
        sys.exit(1)
    return p

def main():
    if len(sys.argv) < 2:
        print("用法: python3 gen_xlsx.py '<JSON>'", file=sys.stderr)
        sys.exit(1)

    data = json.loads(sys.argv[1])
    output_path = safe_output(data["output"])
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 删除默认 sheet

    sheets = data.get("sheets", [])
    if not sheets:
        sheets = [{}]  # 至少保留一个空 sheet，避免 openpyxl 保存无 sheet 的工作簿报错

    for sheet_def in sheets:
        ws = wb.create_sheet(sheet_def.get("name", "Sheet"))
        row_idx = 1

        headers = sheet_def.get("headers", [])
        if headers:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col, value=header)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")
            row_idx += 1

        for row in sheet_def.get("rows", []):
            for col, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col, value=value)
            row_idx += 1

        widths = sheet_def.get("column_widths", [])
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    wb.save(str(output_path))
    print(str(output_path.resolve()))

if __name__ == "__main__":
    main()
